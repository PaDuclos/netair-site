#!/usr/bin/env python3
"""
export_excel.py — Tâche B1/T1 du moteur de prix Netair.

Lit le classeur tarifaire `Calculateur_Netair.xlsx` (la « bibliothèque de grilles »)
et le recopie en données prêtes pour le site :
  - site/src/lib/pricing/data/tables.json       (toutes les grilles)
  - site/src/lib/pricing/data/tables.meta.json  (carte d'identité : source, date, empreinte)

Le moteur de prix (TypeScript) lit UNIQUEMENT tables.json — jamais l'Excel directement.
Ce script est un outil d'atelier : il ne part pas dans le navigateur.

Les cases « Caractéristiques non assurées par le fournisseur » deviennent `null`
(le moteur saura : combinaison non fabriquée).

Usage :
    python3 export_excel.py [chemin_vers_Calculateur_Netair.xlsx]

Sans argument, cherche l'Excel à l'emplacement par défaut (dossier Netair/BLOC1...).
"""

import hashlib
import json
import re
import sys
import unicodedata
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

SENTINEL = "Caractéristiques non assurées par le fournisseur"
CLASSE_RE = re.compile(r"^(G|M|F|E|H|U)\d")  # G4, M5, F7, E10, H13, U15…

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_XLSX = (
    SCRIPT_DIR / "../../../../BLOC1_Gamme_Produits/Grille_Couts_Internes/Calculateur_Netair.xlsx"
).resolve()
OUT_DIR = (SCRIPT_DIR / "../src/lib/pricing/data").resolve()


# ─────────────────────────────────────────────────────────────────────────────
# Petits utilitaires de lecture
# ─────────────────────────────────────────────────────────────────────────────

def clean(v):
    """Valeur de cellule normalisée : sentinel → None, chaîne vidée → None."""
    if v == SENTINEL:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v or None
    return v


def code_str(v):
    """Code gamme toujours en chaîne (cohérent avec la logique INDIRECT de l'Excel)."""
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v)


def header_row(ws):
    return [clean(c.value) for c in ws[1]]


def rows_after_header(ws):
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(c is not None and c != "" for c in row):
            yield row


class _Cols(dict):
    """{en-tête: index} avec message clair si une colonne attendue manque."""

    def __init__(self, headers, sheet):
        super().__init__({h: i for i, h in enumerate(headers) if h})
        self._sheet = sheet

    def __missing__(self, key):
        raise SystemExit(
            f"ERREUR : onglet « {self._sheet} » — colonne attendue introuvable : « {key} ». "
            f"L'Excel a-t-il été modifié ? Colonnes présentes : {sorted(self.keys())}"
        )


def col_map(headers, sheet):
    """{en-tête: index 0-based} pour les colonnes non vides, avec erreur explicite si manquante."""
    return _Cols(headers, sheet)


def classe_cols(headers):
    """Liste (nom_classe, index) des colonnes d'efficacité (G/M/F/E/H/U + chiffre)."""
    return [(h, i) for i, h in enumerate(headers) if h and CLASSE_RE.match(h)]


# ─────────────────────────────────────────────────────────────────────────────
# Lecteurs par onglet
# ─────────────────────────────────────────────────────────────────────────────

def read_prix_grille(ws, dim_cols):
    """
    Onglets Prix_L_et_l / Prix_Surface / Prix_Surface_HF.
    `dim_cols` = liste (clé_json, en-tête Excel) des colonnes de bornes propres à l'onglet.
    Les colonnes de classes sont détectées automatiquement.
    """
    headers = header_row(ws)
    cm = col_map(headers, ws.title)
    classes = classe_cols(headers)
    out = []
    for r in rows_after_header(ws):
        row = {
            "gamme": clean(r[cm["Nom de la gamme"]]),
            "code": code_str(r[cm["Code gamme"]]),
            "ep": clean(r[cm["Epaisseur"]]),
            "qmin": clean(r[cm["Qté mini"]]),
            "qmax": clean(r[cm["Qté maxi"]]),
            "index": clean(r[cm["Index"]]),
            "unite": clean(r[cm["Unité"]]),
        }
        for key, hdr in dim_cols:
            row[key] = clean(r[cm[hdr]])
        row["prix"] = {nom: clean(r[i]) for nom, i in classes}
        out.append(row)
    return out


def read_prix_piece(ws):
    headers = header_row(ws)
    cm = col_map(headers, ws.title)
    out = []
    for r in rows_after_header(ws):
        out.append({
            "gamme": clean(r[cm["Nom de la gamme"]]),
            "code": code_str(r[cm["Code gamme"]]),
            "ep": clean(r[cm["Epaisseur"]]),
            "pd_min": clean(r[cm["Petite dimension mini"]]),
            "pd_max": clean(r[cm["Petite dimension maxi"]]),
            "gd_min": clean(r[cm["Grande dimension mini"]]),
            "gd_max": clean(r[cm["Grande dimension maxi"]]),
            "qmin": clean(r[cm["Qté mini"]]),
            "qmax": clean(r[cm["Qté maxi"]]),
            "index": clean(r[cm["Index"]]),
            "unite": clean(r[cm["Unité"]]),
            "pu": clean(r[cm["PU"]]),
            "commentaire": clean(r[cm["Commentaire"]]),
        })
    return out


def read_poids(ws):
    headers = header_row(ws)
    cm = col_map(headers, ws.title)
    out = []
    for r in rows_after_header(ws):
        out.append({
            "gamme": clean(r[cm["Nom de la gamme"]]),
            "code": code_str(r[cm["Code gamme"]]),
            "ep": clean(r[cm["Epaisseur"]]),
            "index": clean(r[cm["Index"]]),
            "unite": clean(r[cm["Unité"]]),
            "poids": clean(r[cm["Poids"]]),
        })
    return out


def read_expedition(ws):
    headers = header_row(ws)
    cm = col_map(headers, ws.title)
    out = {}
    for r in rows_after_header(ws):
        dep = clean(r[cm["Département"]])
        tarif = clean(r[cm["Tarif expédition"]])
        if dep is not None:
            out[str(dep)] = tarif
    return out


def read_gammes(ws, methodes):
    """`methodes` = {code: lettre} lu dans Infos_Netair. Règle de sécurité : famille
    « hors calculateur » ou méthode inconnue/absente → sur_devis (jamais de prix deviné)."""
    headers = header_row(ws)
    cm = col_map(headers, ws.title)
    remise_cols = [h for h in headers if h and h.startswith("Remise sur catégorie tarif")]
    out = []
    for r in rows_after_header(ws):
        nom = clean(r[cm["Nom de la gamme"]])
        if nom is None:  # ligne « Libre » du calculateur (sur-mesure manuel) → pas un produit boutique
            continue
        code = code_str(r[cm["Code gamme"]])
        famille = clean(r[cm["Famille de gamme"]])
        methode = methodes.get(code)
        if methode is None or (famille and "hors calculateur" in _sans_accents(famille)):
            methode = "sur_devis"
        out.append({
            "famille": famille,
            "code": code,
            "nom": nom,
            "methode": methode,
            "eff_defaut": clean(r[cm["Efficacité : Plus utilisé"]]),
            "ep_defaut": clean(r[cm["Epaisseur : Plus utilisé"]]),
            "coeff": clean(r[cm["Coeff pour ratio prix tarif"]]),
            "ratio": clean(r[cm["Ratio prix tarif"]]),
            "remises": [clean(r[cm[h]]) for h in remise_cols],
            "frais_livraison": clean(r[cm["Frais de livraison"]]),
        })
    return out


def read_params(ws):
    out = []
    for r in rows_after_header(ws):
        libelle = clean(r[0])
        if libelle is None:
            continue
        out.append({
            "libelle": libelle,
            "valeur": clean(r[1]) if len(r) > 1 else None,
            "unite": clean(r[2]) if len(r) > 2 else None,
        })
    return out


def read_iso(ws):
    headers = header_row(ws)
    cm = col_map(headers, ws.title)
    out = {}
    for r in rows_after_header(ws):
        en = clean(r[cm["EN 779"]])
        iso = clean(r[cm["ISO16890"]])
        if en is not None:
            out[str(en)] = iso
    return out


def first_value(ws):
    """Première valeur sous l'en-tête (onglets mono-valeur : franco, durée)."""
    for r in rows_after_header(ws):
        return clean(r[0])
    return None


def _sans_accents(s):
    """Minuscule sans accents — pour comparer des libellés robustement."""
    s = unicodedata.normalize("NFKD", str(s))
    return "".join(c for c in s if not unicodedata.combining(c)).lower().strip()


# Méthode de calcul : 6 méthodes A→F (cf. SPEC_B1 §3) + sur_devis.
#  A = renfort + L×l/surface · B = périmètre · C = surface · D = cadre+média+pièce
#  E = prix pièce · F = lecture L×l directe. Tout libellé inconnu/absent → sur_devis.
def methode_depuis_libelle(libelle):
    """
    Traduit le libellé « Méthode calcul » de l'Excel en lettre de méthode (ordre = priorité).
    Sécurité : tout libellé non reconnu → sur_devis (jamais une fausse méthode). Un libellé
    NON VIDE non reconnu (≠ « Hors calculateur », volontaire) déclenche un avertissement,
    pour qu'une coquille future ne fasse pas disparaître une gamme en silence.
    """
    if libelle is None:
        return "sur_devis"
    t = _sans_accents(libelle)
    if "hors calculateur" in t:
        return "sur_devis"
    if "renfort" in t:
        return "A"
    if "perim" in t:
        return "B"
    if "cadre" in t:
        return "D"
    if "piece" in t:
        return "E"
    if "simple" in t:
        return "F"
    if "surface" in t:
        return "C"
    print(f"AVERTISSEMENT : libellé de méthode non reconnu « {libelle} » → sur_devis. "
          "Vérifier l'orthographe dans Infos_Netair si cette gamme devait être calculable.")
    return "sur_devis"


def read_methodes(ws):
    """
    Lit la correspondance code → méthode dans l'onglet Infos_Netair (tableau semi-libre :
    on repère la ligne d'en-tête contenant « Code » et « Méthode calcul »). Tolérant :
    si l'en-tête est introuvable, renvoie {} (toutes les gammes basculeront en sur_devis).
    """
    rows = list(ws.iter_rows(values_only=True))
    code_idx = meth_idx = header_at = None
    for n, r in enumerate(rows):
        cells = [_sans_accents(c) if c is not None else "" for c in r]
        if "code" in cells and any("methode" in c for c in cells):
            code_idx = cells.index("code")
            meth_idx = next(i for i, c in enumerate(cells) if "methode" in c)
            header_at = n
            break
    if code_idx is None:
        print("AVERTISSEMENT : en-tête « Code / Méthode calcul » introuvable dans Infos_Netair "
              "→ toutes les gammes seront marquées sur_devis.")
        return {}
    out = {}
    for r in rows[header_at + 1:]:
        code = code_str(clean(r[code_idx])) if code_idx < len(r) else None
        libelle = clean(r[meth_idx]) if meth_idx < len(r) else None
        if code is None or libelle is None:  # ligne sans code ou sans méthode → ignorée
            continue
        out[code] = methode_depuis_libelle(libelle)
    return out


# ─────────────────────────────────────────────────────────────────────────────
# Programme principal
# ─────────────────────────────────────────────────────────────────────────────

def main():
    xlsx = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx.exists():
        sys.exit(f"ERREUR : Excel introuvable : {xlsx}")

    wb = load_workbook(xlsx, data_only=True)

    requis = [
        "Prix_L_et_l", "Prix_Surface", "Prix_Surface_HF", "Prix_Pièce", "Poids_filtres",
        "Tableau_Tarifs_Expédition", "Tableau_franco", "Tableau_Gammes",
        "Paramètres unitaires", "ISO_16890", "Durée_validité",
    ]
    manquants = [s for s in requis if s not in wb.sheetnames]
    if manquants:
        sys.exit(
            "ERREUR : onglet(s) attendu(s) introuvable(s) dans l'Excel : "
            f"{manquants}. Onglets présents : {wb.sheetnames}"
        )

    methodes = read_methodes(wb["Infos_Netair"]) if "Infos_Netair" in wb.sheetnames else {}

    tables = {
        "prix_l_et_l": read_prix_grille(wb["Prix_L_et_l"], [
            ("pd_min", "Petite dimension mini"), ("pd_max", "Petite dimension maxi"),
            ("gd_min", "Grande dimension mini"), ("gd_max", "Grande dimension maxi"),
        ]),
        "prix_surface": read_prix_grille(wb["Prix_Surface"], [
            ("surf_min", "Surface mini"), ("surf_max", "Surface maxi"),
        ]),
        "prix_surface_hf": read_prix_grille(wb["Prix_Surface_HF"], []),
        "prix_piece": read_prix_piece(wb["Prix_Pièce"]),
        "poids": read_poids(wb["Poids_filtres"]),
        "expedition": read_expedition(wb["Tableau_Tarifs_Expédition"]),
        "franco": first_value(wb["Tableau_franco"]),
        "gammes": read_gammes(wb["Tableau_Gammes"], methodes),
        "params": read_params(wb["Paramètres unitaires"]),
        "iso16890": read_iso(wb["ISO_16890"]),
        "validite_jours": first_value(wb["Durée_validité"]),
    }

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    tables_json = json.dumps(tables, ensure_ascii=False, indent=2, sort_keys=True)
    (OUT_DIR / "tables.json").write_text(tables_json + "\n", encoding="utf-8")

    # Version interne déclarée dans l'onglet Infos_Netair (ligne « Version »), si présente.
    version_excel = None
    if "Infos_Netair" in wb.sheetnames:
        for r in wb["Infos_Netair"].iter_rows(values_only=True):
            if r and r[0] == "Version" and len(r) > 1:
                version_excel = clean(r[1])
                break

    excel_sha = hashlib.sha256(xlsx.read_bytes()).hexdigest()
    data_sha = hashlib.sha256(tables_json.encode("utf-8")).hexdigest()
    meta = {
        "source": xlsx.name,
        "version_excel": version_excel,
        "exporte_le": date.today().isoformat(),
        "excel_sha256": excel_sha,
        "tables_sha256": data_sha,
        "version_outil": "1.0",
        "lignes_par_table": {
            k: (len(v) if isinstance(v, (list, dict)) else 1) for k, v in tables.items()
        },
    }
    (OUT_DIR / "tables.meta.json").write_text(
        json.dumps(meta, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )

    print(f"OK — Excel lu : {xlsx.name}")
    print(f"     → {OUT_DIR / 'tables.json'}")
    print(f"     → {OUT_DIR / 'tables.meta.json'}")
    print("Lignes par table :")
    for k, n in meta["lignes_par_table"].items():
        print(f"  - {k:18s} : {n}")

    print("Aiguillage des gammes (code → méthode) :")
    for g in tables["gammes"]:
        print(f"  - {str(g['code']):>4} {str(g['nom'])[:24]:24s} → {g['methode']}")


if __name__ == "__main__":
    main()
