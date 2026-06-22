#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Vérification de cohérence : Bibliothèque des fiches techniques  <->  fiches HTML.

Pour chaque référence de la bibliothèque, le script :
  1. cherche la fiche HTML correspondante dans ce dossier
     (nom attendu : "Fiche technique <REF> v<N>.html", la version la plus haute est retenue) ;
  2. extrait les données clés de la fiche (EN 779, ISO 16890, version) ;
  3. les compare aux colonnes de la bibliothèque et signale les écarts.

Usage :
    python3 verif_coherence.py
"""
import os, re, html, glob, sys
from openpyxl import load_workbook

DOSSIER = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(DOSSIER, "Bibliotheque_Fiches_Techniques_Netair.xlsx")

# Colonnes de la bibliothèque (1-based) — en-tête sur la ligne 3, données dès la ligne 4
COL = {"num": 1, "ref": 2, "en779": 5, "en16890": 6, "version": 7, "fichier": 10}
LIGNE_ENTETE = 3


def norm(txt):
    """Normalise pour comparaison : minuscules, flèches/espaces unifiés."""
    if txt is None:
        return ""
    t = str(txt).strip().lower()
    t = t.replace("→", ">").replace("->", ">").replace(" à ", " > ")
    t = re.sub(r"\s+", " ", t)
    t = t.replace(" > ", ">").replace("> ", ">").replace(" >", ">")
    return t


def lire_fiche(chemin):
    """Extrait {en779, en16890, version_fichier} d'une fiche HTML bundle."""
    raw = open(chemin, encoding="utf-8", errors="ignore").read()
    raw = raw.replace("\\u002F", "/").replace("\\u003C", "<").replace("\\u003E", ">")
    raw = raw.replace("\\n", " ")
    raw = html.unescape(raw)

    def cherche(motif):
        m = re.search(motif, raw)
        return m.group(1).strip() if m else None

    en779 = cherche(r"EN\s*779\s*:\s*([^<]+?)\s*<")
    en16890 = cherche(r"ISO\s*16890\s*:\s*([^<]+?)\s*<")
    mver = re.search(r" v(\d+)\.html$", os.path.basename(chemin))
    version = "v" + mver.group(1) if mver else None
    return {"en779": en779, "en16890": en16890, "version": version}


def trouver_fiche(ref):
    """Retourne le chemin de la fiche la plus récente pour une référence, ou None."""
    motif = os.path.join(DOSSIER, f"Fiche technique {ref} v*.html")
    fichiers = glob.glob(motif)
    if not fichiers:
        return None

    def num(p):
        m = re.search(r" v(\d+)\.html$", p)
        return int(m.group(1)) if m else 0

    return max(fichiers, key=num)


def main():
    if not os.path.exists(XLSX):
        sys.exit(f"Introuvable : {XLSX}")
    wb = load_workbook(XLSX)
    ws = wb.active

    total = ecarts = sans_fiche = ok = 0
    rapport = []

    for row in ws.iter_rows(min_row=LIGNE_ENTETE + 1, max_row=ws.max_row):
        ref = row[COL["ref"] - 1].value
        if not ref:
            continue
        total += 1
        chemin = trouver_fiche(ref)
        if not chemin:
            sans_fiche += 1
            rapport.append(("—", ref, "Aucune fiche HTML (À créer)", []))
            continue

        fiche = lire_fiche(chemin)
        biblio = {
            "en779": row[COL["en779"] - 1].value,
            "en16890": row[COL["en16890"] - 1].value,
            "version": row[COL["version"] - 1].value,
        }
        problemes = []
        for champ, libelle in [("en779", "EN 779"), ("en16890", "ISO 16890"), ("version", "Version")]:
            v_fiche, v_biblio = fiche.get(champ), biblio.get(champ)
            if v_fiche is None:
                continue
            if norm(v_fiche) != norm(v_biblio):
                problemes.append(f"{libelle} : biblio «{v_biblio}»  ≠  fiche «{v_fiche}»")

        if problemes:
            ecarts += 1
            rapport.append(("✗", ref, os.path.basename(chemin), problemes))
        else:
            ok += 1
            rapport.append(("✓", ref, os.path.basename(chemin), []))

    print("=" * 70)
    print("  VÉRIFICATION DE COHÉRENCE — Bibliothèque ↔ Fiches techniques")
    print("=" * 70)
    for statut, ref, fich, problemes in rapport:
        print(f"\n[{statut}] {ref}   ({fich})")
        for p in problemes:
            print(f"      ⚠  {p}")
    print("\n" + "-" * 70)
    print(f"  {total} références   |   ✓ {ok} cohérentes   |   "
          f"✗ {ecarts} en écart   |   — {sans_fiche} sans fiche")
    print("-" * 70)
    return 1 if ecarts else 0


if __name__ == "__main__":
    sys.exit(main())
