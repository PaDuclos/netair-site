#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Construit DONNEES_PDC_Netair.xlsx — base de tests R&D perte de charge."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DOSSIER = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(DOSSIER, "DONNEES_PDC_Netair.xlsx")
FONT = "Arial"

DEBITS = [1000, 1500, 2000, 2500, 3000, 3500, 4000]   # m³/h (section 592×592)

# N° fiche, Famille, Référence, Classe EN779, ISO 16890, Épaisseur, Dimensions
ROWS = [
 ("FT-NETMETAL","Préfiltres","NETMETAL","G1 / G2","ISO Coarse","","592×592×__"),
 ("FT-NETFIL","Préfiltres","NETFIL","G2 / G3","ISO Coarse","","592×592×__"),
 ("FT-NETFIBRE","Préfiltres","NETFIBRE","G2 → M5","ISO Coarse → ePM10","","rouleau"),
 ("FT-NETPLY","Préfiltres","NETPLY","G4","Coarse 65%","48","592×592×48"),
 ("FT-NETPLY","Préfiltres","NETPLY","G4","Coarse 65%","98","592×592×98"),
 ("FT-NETPLY","Préfiltres","NETPLY","M5","ePM10 50%","48","592×592×48"),
 ("FT-NETPLY","Préfiltres","NETPLY","M5","ePM10 50%","98","592×592×98"),
 ("FT-NETPLAN","Préfiltres","NETPLAN","G3 / G4","ISO Coarse","","592×592×__"),
 ("FT-NETBAG-S","Poches souples","NETBAG S","G4 → F9","ePM10 → ePM1","","592×592×380"),
 ("FT-NETPAK-S-BORA","Compacts / poches rigides","NETPAK S BORA","M5 → F9","ePM10 → ePM1","","592×592×292"),
 ("FT-NETPAK-S-CILIA","Compacts / poches rigides","NETPAK S CILIA","M5 → F9","ePM1 55% → 70%","","592×592×__"),
 ("FT-NETPAK-S-AZUR","Compacts / poches rigides","NETPAK S AZUR","M5 → F9","ePM10 → ePM1","","592×592×292"),
 ("FT-NETPAK-S-LUMEN","Compacts / poches rigides","NETPAK S LUMEN","M5 → F9","ePM10 → ePM1","","592×592×292"),
 ("FT-NETCEL-V-AZUR","HEPA / T.H.E","NETCEL V AZUR","E10 / H13","ISO 50 U / 35 H","","592×592×292"),
 ("FT-NETCEL-V-NIVAL","HEPA / T.H.E","NETCEL V NIVAL","E10 → H14","ISO 35 H","","592×592×292"),
 ("FT-NETPAK-V-LAM","HEPA / T.H.E","NETPAK V LAM","H14","ISO 15 U","","610×610×__"),
 ("FT-NETCARB-CILIA","Charbon actif","NETCARB CILIA","—","Moléculaire","","592×592×48"),
 ("FT-NETCARB-AZUR","Charbon actif","NETCARB AZUR","—","Moléculaire","","592×592×292"),
 ("FT-NETCARB-NIVAL","Charbon actif","NETCARB NIVAL","—","Moléculaire","","610×610×292"),
 ("FT-NETCARB-BAG","Charbon actif","NETCARB BAG","—","Moléculaire","","592×592×380"),
 ("FT-NETPAK-S-DUO","Combinés","NETPAK S DUO","F7 + CA","ePM1 55% + Mol.","","592×592×98"),
]

wb = Workbook()
ws = wb.active
ws.title = "Données PDC R&D"

BLUE = "1F3864"; TEAL = "0897A5"
hdr_fill = PatternFill("solid", fgColor=BLUE)
in_fill = PatternFill("solid", fgColor="FFFDE7")   # saisie ΔP
calc_fill = PatternFill("solid", fgColor="EEF2F8")  # calculé
param_fill = PatternFill("solid", fgColor="FFF2CC")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
C = Alignment(horizontal="center", vertical="center", wrap_text=True)
L = Alignment(horizontal="left", vertical="center", wrap_text=True)
R = Alignment(horizontal="right", vertical="center")

# colonnes : A..G identif | H..N ΔP(7) | O a | P b | Q c | R R² | S poly | T pts | U ΔP nom | V date
NID = 7                 # colonnes identification (A-G)
NDP = len(DEBITS)       # 7 colonnes ΔP -> H(8)..N(14)
first_dp = NID + 1      # 8 -> H
last_dp = NID + NDP     # 14 -> N
col_a = last_dp + 1     # O
col_b = col_a + 1; col_c = col_b + 1; col_r2 = col_c + 1
col_poly = col_r2 + 1; col_pts = col_poly + 1; col_nom = col_pts + 1; col_date = col_nom + 1
NCOL = col_date

def cl(i): return get_column_letter(i)
DP1, DPN = cl(first_dp), cl(last_dp)

# Titre
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOL)
ws.cell(1,1,"NETAIR — Données R&D · Perte de charge (ΔP)").font = Font(name=FONT,bold=True,size=14,color=BLUE)
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NCOL)
ws.cell(2,1,"Modèle  ΔP = a·v² + b·v + c  —  saisir les ΔP mesurées (Pa, cases jaunes) ; vitesse, polynôme et R² se calculent automatiquement.").font = Font(name=FONT,italic=True,size=9,color="808080")

# Paramètres (ligne 3)
ws.cell(3,1,"Section frontale (m²) :").font = Font(name=FONT,bold=True,size=9)
ws.merge_cells("A3:B3")
pc = ws.cell(3,3,"=0.592*0.592"); pc.font=Font(name=FONT,bold=True,size=9); pc.fill=param_fill; pc.border=border; pc.number_format="0.0000"; pc.alignment=C
ws.cell(3,5,"Débit nominal fiche (m³/h) :").font = Font(name=FONT,bold=True,size=9)
ws.merge_cells("E3:G3")
hc = ws.cell(3,8,3400); hc.font=Font(name=FONT,bold=True,size=9); hc.fill=param_fill; hc.border=border; hc.number_format="0"; hc.alignment=C

# En-tête (ligne 5) + débits
H = 5
labels = {1:"N° FICHE",2:"FAMILLE",3:"RÉFÉRENCE",4:"CLASSE\nEN 779",5:"ISO 16890",6:"ÉP.\n(mm)",7:"DIMENSIONS"}
for i,t in labels.items():
    c = ws.cell(H,i,t); c.fill=hdr_fill; c.font=Font(name=FONT,bold=True,color="FFFFFF",size=9); c.alignment=C; c.border=border
for k,d in enumerate(DEBITS):
    c = ws.cell(H, first_dp+k, d); c.fill=hdr_fill; c.font=Font(name=FONT,bold=True,color="FFFFFF",size=9); c.alignment=C; c.border=border; c.number_format="0"
for i,t in {col_a:"a\n(·v²)",col_b:"b\n(·v)",col_c:"c",col_r2:"R²",col_poly:"POLYNÔME  (a·v²+b·v+c)",col_pts:"Pts",col_nom:"ΔP nom.\n(Pa)",col_date:"DATE\nTEST"}.items():
    c = ws.cell(H,i,t); c.fill=hdr_fill; c.font=Font(name=FONT,bold=True,color="FFFFFF",size=9); c.alignment=C; c.border=border

# Ligne vitesse (6)
V = 6
vc = ws.cell(V,7,"Vitesse (m/s) →"); vc.font=Font(name=FONT,italic=True,size=9,color=TEAL); vc.alignment=R
for k,d in enumerate(DEBITS):
    col = first_dp+k
    c = ws.cell(V,col,f"={cl(col)}{H}/3600/$C$3"); c.font=Font(name=FONT,italic=True,size=9,color=TEAL); c.alignment=C; c.border=border; c.number_format="0.00"
ws.cell(V,col_poly,"Vitesse déduite de la section 592×592").font=Font(name=FONT,italic=True,size=8,color="A0A0A0")

# Données (à partir de la ligne 7)
start = 7
for ridx,row in enumerate(ROWS):
    r = start + ridx
    for ci,val in enumerate(row, start=1):
        c = ws.cell(r,ci,val); c.font=Font(name=FONT,size=9); c.border=border
        c.alignment = L if ci in (2,3,7) else C
    ws.cell(r,1).font = Font(name=FONT,size=9,bold=True,color=BLUE)
    ws.cell(r,3).font = Font(name=FONT,size=9,bold=True)
    # ΔP saisie (vides)
    for k in range(NDP):
        c = ws.cell(r, first_dp+k); c.fill=in_fill; c.border=border; c.alignment=C; c.number_format="0"
    rng_y = f"$H{r}:$N{r}"
    linest = f"LINEST({rng_y},$H$6:$N$6^{{1;2}})"
    ws.cell(r,col_a, f'=IFERROR(IF(COUNT({rng_y})=7,INDEX({linest},1,1),""),"")')
    ws.cell(r,col_b, f'=IFERROR(IF(COUNT({rng_y})=7,INDEX({linest},1,2),""),"")')
    ws.cell(r,col_c, f'=IFERROR(IF(COUNT({rng_y})=7,INDEX({linest},1,3),""),"")')
    ws.cell(r,col_r2,f'=IFERROR(IF(COUNT({rng_y})=7,INDEX(LINEST({rng_y},$H$6:$N$6^{{1;2}},TRUE,TRUE),3,1),""),"")')
    ws.cell(r,col_poly,f'=IF({cl(col_a)}{r}="","",TEXT({cl(col_a)}{r},"0.00")&"·v² "&IF({cl(col_b)}{r}>=0,"+ ","− ")&TEXT(ABS({cl(col_b)}{r}),"0.00")&"·v "&IF({cl(col_c)}{r}>=0,"+ ","− ")&TEXT(ABS({cl(col_c)}{r}),"0.00"))')
    ws.cell(r,col_pts,f"=COUNT({rng_y})")
    ws.cell(r,col_nom,f'=IFERROR(IF({cl(col_a)}{r}="","",{cl(col_a)}{r}*($H$3/3600/$C$3)^2+{cl(col_b)}{r}*($H$3/3600/$C$3)+{cl(col_c)}{r}),"")')
    for ci in (col_a,col_b,col_c,col_r2,col_poly,col_pts,col_nom,col_date):
        c = ws.cell(r,ci); c.border=border; c.alignment=(L if ci==col_poly else C); c.font=Font(name=FONT,size=9)
        if ci in (col_a,col_b,col_c): c.number_format="0.00"; c.fill=calc_fill
        elif ci==col_r2: c.number_format="0.000"; c.fill=calc_fill
        elif ci in (col_pts,col_nom): c.fill=calc_fill; c.number_format=("0" if ci==col_nom else "0")
        elif ci==col_poly: c.fill=calc_fill; c.font=Font(name=FONT,size=9,bold=True,color=BLUE)
last = start + len(ROWS) - 1

# largeurs
widths = {"A":17,"B":15,"C":15,"D":10,"E":13,"F":7,"G":13}
for k in range(NDP): widths[cl(first_dp+k)] = 7
widths[cl(col_a)]=8; widths[cl(col_b)]=8; widths[cl(col_c)]=8; widths[cl(col_r2)]=8
widths[cl(col_poly)]=26; widths[cl(col_pts)]=5; widths[cl(col_nom)]=9; widths[cl(col_date)]=12
for col,w in widths.items(): ws.column_dimensions[col].width = w
ws.row_dimensions[H].height = 30

ws.freeze_panes = "H7"
wb.save(OUT)
print("OK", OUT, "| lignes:", len(ROWS))
EOF = None
