#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Lecture des coefficients polynomiaux (a,b,c) depuis DONNEES_PDC_Netair.xlsx."""
import os
from openpyxl import load_workbook

DOSSIER = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(DOSSIER, "DONNEES_PDC_Netair.xlsx")
COL = {"ref": 3, "classe": 4, "ep": 6, "a": 15, "b": 16, "c": 17}
START = 7

def _eff(classe):
    c = (classe or "").upper()
    if "G4" in c or "COARSE" in c: return "g4"
    if "M5" in c: return "m5"
    if "M6" in c: return "m6"
    return None

def lire(ref_voulue):
    """Retourne {eff: {ep(int): (a,b,c)}} pour une référence donnée (lignes renseignées)."""
    ws = load_workbook(XLSX, data_only=True).active
    out = {}
    for r in range(START, ws.max_row + 1):
        ref = ws.cell(r, COL["ref"]).value
        a = ws.cell(r, COL["a"]).value
        if ref != ref_voulue or not isinstance(a, (int, float)):
            continue
        eff = _eff(ws.cell(r, COL["classe"]).value)
        try:
            ep = int(str(ws.cell(r, COL["ep"]).value).strip())
        except (ValueError, AttributeError):
            continue
        if eff is None:
            continue
        b = ws.cell(r, COL["b"]).value
        c = ws.cell(r, COL["c"]).value
        out.setdefault(eff, {})[ep] = (round(a, 4), round(b, 4), round(c, 4))
    return out

def poly_block_netply():
    """Bloc JS (avec \\n littéraux) du POLY NETPLY pour insertion dans la fiche."""
    d = lire("NETPLY")
    manquant = [k for k in (("g4", 48), ("g4", 98), ("m5", 48), ("m5", 98))
                if k[1] not in d.get(k[0], {})]
    if manquant:
        raise SystemExit(f"Coefficients NETPLY manquants dans DONNEES_PDC : {manquant} "
                         f"(renseigne les 7 points de chaque variante).")
    def var(eff, ep):
        a, b, c = d[eff][ep]
        return f"{ep}: {{a:{a},b:{b},c:{c}}}"
    g4 = f"      g4: {{ {var('g4',48)}, {var('g4',98)} }},\\n"
    m5 = f"      m5: {{ {var('m5',48)}, {var('m5',98)} }},\\n"
    return g4 + m5

if __name__ == "__main__":
    print(poly_block_netply().replace("\\n", "\n"))
