#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Génère 'index.html' — sommaire des fiches techniques — à partir de la
bibliothèque Excel, pour une consultation fluide dans le navigateur
(liens vers les fiches, sans blocage sandbox d'Excel).

Usage :
    python3 generer_index.py
"""
import os, re, glob, html
from openpyxl import load_workbook

DOSSIER = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(DOSSIER, "Bibliotheque_Fiches_Techniques_Netair.xlsx")
SORTIE = os.path.join(DOSSIER, "index.html")
LIGNE_ENTETE = 3
COL = {"num": 0, "ref": 1, "fam": 2, "type": 3, "en779": 4,
       "en16890": 5, "version": 6, "statut": 8}


def trouver_fiche(ref):
    fichiers = glob.glob(os.path.join(DOSSIER, f"Fiche technique {ref} v*.html"))
    if not fichiers:
        return None
    return os.path.basename(max(fichiers, key=lambda p: int(re.search(r" v(\d+)\.html$", p).group(1))))


def esc(v):
    return html.escape(str(v)) if v is not None else ""


def main():
    wb = load_workbook(XLSX)
    ws = wb.active

    familles = {}
    n_total = n_dispo = 0
    for row in ws.iter_rows(min_row=LIGNE_ENTETE + 1, max_row=ws.max_row):
        ref = row[COL["ref"]].value
        if not ref:
            continue
        n_total += 1
        fiche = trouver_fiche(ref)
        if fiche:
            n_dispo += 1
        item = {
            "num": row[COL["num"]].value,
            "ref": ref,
            "type": row[COL["type"]].value,
            "en779": row[COL["en779"]].value,
            "en16890": row[COL["en16890"]].value,
            "version": row[COL["version"]].value,
            "statut": row[COL["statut"]].value,
            "fiche": fiche,
        }
        familles.setdefault(row[COL["fam"]].value, []).append(item)

    sections = []
    for fam, items in familles.items():
        lignes = []
        for it in items:
            dispo = it["fiche"] is not None
            if dispo:
                num_html = (f'<a class="num" href="{esc(it["fiche"])}" target="_blank" '
                            f'rel="noopener">{esc(it["num"])}</a>')
            else:
                num_html = f'<span class="num off">{esc(it["num"])}</span>'
            st = (it["statut"] or "").strip()
            cls = {"Validée": "ok", "À créer": "todo",
                   "Brouillon": "draft", "À réviser": "rev"}.get(st, "todo")
            ver = esc(it["version"]) if it["version"] else "—"
            lignes.append(f"""        <tr>
          <td>{num_html}</td>
          <td class="ref">{esc(it["ref"])}</td>
          <td>{esc(it["type"])}</td>
          <td class="c">{esc(it["en779"])}</td>
          <td class="c">{esc(it["en16890"])}</td>
          <td class="c">{ver}</td>
          <td class="c"><span class="badge {cls}">{esc(st)}</span></td>
        </tr>""")
        sections.append(f"""    <section>
      <h2>{esc(fam)}</h2>
      <table>
        <thead><tr>
          <th>N° fiche</th><th>Référence</th><th>Type de produit</th>
          <th>EN 779</th><th>ISO 16890</th><th>Version</th><th>Statut</th>
        </tr></thead>
        <tbody>
{chr(10).join(lignes)}
        </tbody>
      </table>
    </section>""")

    doc = f"""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>NETAIR — Bibliothèque des fiches techniques</title>
<style>
  :root {{ --blue:#0F3261; --teal:#0897A5; --line:#E3E9F2; --bg:#F7F9FC; --ink:#3a4654; --mut:#7A8696; }}
  * {{ box-sizing:border-box; }}
  body {{ margin:0; font-family:'Helvetica Neue',Arial,sans-serif; color:var(--ink); background:#fff; }}
  header {{ padding:36px 40px 26px; border-bottom:3px solid var(--blue); }}
  h1 {{ margin:0; font-size:26px; color:var(--blue); letter-spacing:.3px; }}
  .sub {{ margin-top:6px; color:var(--mut); font-size:13px; }}
  .stats {{ margin-top:14px; display:flex; gap:10px; flex-wrap:wrap; }}
  .pill {{ font-size:12px; padding:5px 12px; border-radius:20px; background:var(--bg); color:var(--blue); border:1px solid var(--line); }}
  main {{ padding:10px 40px 60px; max-width:1200px; }}
  section {{ margin-top:34px; }}
  h2 {{ font-size:13px; text-transform:uppercase; letter-spacing:1.4px; color:var(--blue);
        border-left:4px solid var(--teal); padding-left:10px; margin:0 0 12px; }}
  table {{ width:100%; border-collapse:collapse; font-size:13px; }}
  thead th {{ text-align:left; background:var(--blue); color:#fff; font-weight:600; padding:9px 12px; font-size:11px;
              text-transform:uppercase; letter-spacing:.4px; }}
  tbody td {{ padding:9px 12px; border-bottom:1px solid var(--line); }}
  tbody tr:hover {{ background:var(--bg); }}
  td.c {{ text-align:center; white-space:nowrap; }}
  .ref {{ font-weight:700; color:var(--blue); }}
  .num {{ font-family:'IBM Plex Mono','Courier New',monospace; font-weight:700; font-size:12px;
          color:var(--teal); text-decoration:none; }}
  a.num:hover {{ text-decoration:underline; }}
  .num.off {{ color:#B7C0CE; }}
  .badge {{ font-size:11px; font-weight:600; padding:3px 10px; border-radius:20px; }}
  .badge.ok {{ background:#E4F4E8; color:#1B7A3D; }}
  .badge.todo {{ background:#FCEBDD; color:#9C5700; }}
  .badge.draft {{ background:#FFF6DA; color:#8A6D00; }}
  .badge.rev {{ background:#FFE9C2; color:#8A5A00; }}
  footer {{ padding:20px 40px 40px; color:var(--mut); font-size:11px; }}
</style>
</head>
<body>
<header>
  <h1>NETAIR — Bibliothèque des fiches techniques</h1>
  <div class="sub">Cliquez sur un N° de fiche disponible pour ouvrir sa dernière version.</div>
  <div class="stats">
    <span class="pill">{n_total} références</span>
    <span class="pill">{n_dispo} fiche(s) disponible(s)</span>
    <span class="pill">{n_total - n_dispo} à créer</span>
  </div>
</header>
<main>
{chr(10).join(sections)}
</main>
<footer>Généré depuis Bibliotheque_Fiches_Techniques_Netair.xlsx · NETAIR</footer>
</body>
</html>"""

    open(SORTIE, "w", encoding="utf-8").write(doc)
    print(f"index.html généré — {n_total} références, {n_dispo} fiche(s) liée(s)")


if __name__ == "__main__":
    main()
