#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
maj_fiches.py — Pipeline de données « perte de charge ».

Source de vérité UNIQUE : ../DONNEES_PDC_Netair.xlsx (cases ΔP par débit, R&D).
Le script :
  1. lit les ΔP mesurées de chaque ligne (débits 1000→4000) ;
  2. recalcule le polynôme ΔP = a·v² + b·v + c (numpy — indépendant du recalcul Excel) ;
  3. met à jour les coefficients (et les ΔP de référence) dans produits/<slug>.json ;
  4. si la ΔP d'une fiche a changé → incrémente la version (v1.x) + date du jour ;
  5. régénère les fiches modifiées (generer.py → courbe + calculateur + synchro site).

Usage :
  python3 maj_fiches.py            # APERÇU (dry-run) : montre ce qui changerait, n'écrit rien
  python3 maj_fiches.py --apply    # applique : écrit les JSON, bump version, régénère, synchro site

Règle d'or : une ligne Excel SANS ΔP brutes (ex. NETCEL, saisi en coefficients) est IGNORÉE
(le JSON n'est jamais écrasé faute de données). Aucune valeur inventée.
"""

import openpyxl
import numpy as np
import json
import os
import re
import sys
import glob
import subprocess

ROOT = os.path.dirname(os.path.abspath(__file__))
PDC = os.path.abspath(os.path.join(ROOT, "..", "DONNEES_PDC_Netair.xlsx"))
PRODUITS = os.path.join(ROOT, "produits")
AREA = 0.592 * 0.592               # section frontale de référence (grille de vitesse DONNEES_PDC)
APPLY = "--apply" in sys.argv


# --------------------------------------------------------------- helpers ----
def norm_class(v):
    """'M5'->'m5' · 'H13 (EN 1822)'->'h13' · 'E10 ...'->'e10' · '—'/None->'' """
    if not v:
        return ""
    tok = str(v).strip().split()[0]
    return "" if tok in ("—", "-", "") else tok.lower()


def parse_len(ep, dim):
    """Longueur/épaisseur de la ligne : colonne ÉP si numérique, sinon le P de 'L×H×P'."""
    if isinstance(ep, (int, float)):
        return int(ep)
    if dim:
        m = re.findall(r"\d+", str(dim))
        if m:
            return int(m[-1])
    return None


def fit_poly(velocities, dps):
    a, b, c = np.polyfit(np.array(velocities), np.array(dps), 2)
    return round(float(a), 3), round(float(b), 3), round(float(c), 3)


def poly_at(co, v):
    return co[0] * v * v + co[1] * v + co[2]


def numfiche_base(num):
    """'FT-NETCARB-CILIA-001' -> 'FT-NETCARB-CILIA' (retire le suffixe -NNN)."""
    return re.sub(r"-\d+$", "", num)


# ------------------------------------------------------ lecture DONNEES_PDC --
def read_pdc():
    wb = openpyxl.load_workbook(PDC, data_only=True)
    ws = wb.active
    debits = [ws.cell(5, c).value for c in range(8, 15)]          # H5..N5
    velocities_all = [d / 3600 / AREA for d in debits]
    lookup = {}                                                   # (numfiche, classe, longueur) -> {coef, npts}
    for r in range(6, ws.max_row + 1):
        num = ws.cell(r, 1).value
        if not num or not str(num).startswith("FT-"):
            continue
        # On ne retient que les lignes du cadre STANDARD 592×592 : ce sont elles que
        # représentent les courbes des fiches (en vitesse). Les variantes 490/287 ont une
        # vitesse différente à débit égal → gérées par le tableau dimensions, pas une courbe.
        dim = ws.cell(r, 7).value
        nums_dim = re.findall(r"\d+", str(dim or ""))
        if len(nums_dim) >= 2 and not (nums_dim[0] == "592" and nums_dim[1] == "592"):
            continue
        dps_raw = [ws.cell(r, c).value for c in range(8, 15)]
        pts = [(v, p) for v, p in zip(velocities_all, dps_raw) if isinstance(p, (int, float))]
        if len(pts) < 3:                                          # pas assez de points -> ligne ignorée
            continue
        vs = [p[0] for p in pts]
        ds = [p[1] for p in pts]
        coef = fit_poly(vs, ds)
        key = (str(num).strip(), norm_class(ws.cell(r, 4).value),
               parse_len(ws.cell(r, 6).value, ws.cell(r, 7).value))
        lookup[key] = {"coef": coef, "npts": len(pts)}
    return lookup


def find(lookup, num, cls, length):
    """Cherche une ligne par (num, classe, longueur), avec tolérance sur classe/longueur."""
    if (num, cls, length) in lookup:
        return lookup[(num, cls, length)]
    # tolérance : même num + même longueur (classe vide côté charbon)
    for (n, c, l), val in lookup.items():
        if n == num and l == length and (c == cls or c == "" or cls == ""):
            return val
    # mono : même num, une seule ligne
    rows = [v for (n, c, l), v in lookup.items() if n == num]
    return rows[0] if len(rows) == 1 else None


# --------------------------------------------------------- mise à jour JSON --
def co_dict(coef):
    return {"a": coef[0], "b": coef[1], "c": coef[2]}


def same(co_json, coef, vmax=3.17, tol=1.0):
    """Vrai si les deux courbes ΔP coïncident à < tol (Pa) sur la plage RÉELLE du produit
    [0,5 ; vmax] : on ignore le bruit d'arrondi (3e décimale) et l'extrapolation hors plage,
    on ne signale qu'un VRAI changement de perte de charge dans le domaine d'usage."""
    ao, bo, c0 = co_json.get("a", 0), co_json.get("b", 0), co_json.get("c", 0)
    an, bn, cn = coef
    vs = [0.5 + 0.25 * i for i in range(int((vmax - 0.5) / 0.25) + 1)] + [vmax]
    dev = max(abs((ao - an) * v * v + (bo - bn) * v + (c0 - cn)) for v in vs)
    return dev < tol


def update_one(path, lookup):
    """Retourne (liste_de_changements, dict_json_modifié|None)."""
    d = json.load(open(path, encoding="utf-8"))
    num = numfiche_base(d.get("fiche", {}).get("num", ""))
    vmax = d.get("vmax", 3.17)
    changes = []

    def set_poly_slot(poly_obj, ep, coef, tag):
        cur = poly_obj.get(ep, {})
        if same(cur, coef, vmax):                 # pas de vrai changement → on ne touche à rien
            return
        changes.append(f"{tag} ép{ep}: {cur.get('a')}/{cur.get('b')}/{cur.get('c')} → "
                       f"{coef[0]}/{coef[1]}/{coef[2]}")
        poly_obj[ep] = co_dict(coef)

    # --- mode SÉRIE (courbes[] : cls + len + a/b/c)
    if d.get("series"):
        for s in d.get("courbes", []):
            row = find(lookup, num, s.get("cls", ""), s.get("len"))
            if not row:
                continue
            co = row["coef"]
            if same(s, co, vmax):                  # pas de vrai changement → on ne touche pas
                continue
            changes.append(f"{s.get('cls')} {s.get('len')}mm: "
                           f"{s.get('a')}/{s.get('b')}/{s.get('c')} → {co[0]}/{co[1]}/{co[2]}")
            s["a"], s["b"], s["c"] = co
            vnom = d.get("debit_nom", 3400) / 3600 / AREA
            s["dp"] = round(poly_at(co, vnom))

    # --- mode MULTI (classes_list[] : id + poly[ep] + points[ep])
    elif d.get("multi_classe"):
        for cl in d.get("classes_list", []):
            for ep in list(cl.get("poly", {}).keys()):
                row = find(lookup, num, cl.get("id", ""), int(ep))
                if not row:
                    continue
                set_poly_slot(cl["poly"], ep, row["coef"], f"{cl.get('label')}")

    # --- mode LEGACY / MONO / DEUX_ÉPAISSEURS (classes.low/high.poly[48|98])
    elif "classes" in d:
        low, high = d["classes"].get("low"), d["classes"].get("high")
        if d.get("deux_epaisseurs"):                              # 2 courbes réelles 48 / 98
            for ep in ("48", "98"):
                row = find(lookup, num, "", int(ep))
                if not row:
                    continue
                set_poly_slot(low["poly"], ep, row["coef"], f"{low.get('label')}")
                if high:
                    high["poly"][ep] = co_dict(row["coef"])
        elif d.get("mono_classe"):                                # 1 courbe -> remplit 48 et 98
            row = find(lookup, num, norm_class(low.get("label")), None)
            if row:
                for ep in ("48", "98"):
                    set_poly_slot(low["poly"], ep, row["coef"], f"{low.get('label')}")
                    if high:
                        high["poly"][ep] = co_dict(row["coef"])
        # (legacy 2 classes type NETPLY : absent de DONNEES_PDC -> rien)

    if not changes:
        return [], None

    # --- recalcul des ΔP de référence affichées (cohérence courbe ↔ tableau)
    recompute_dp(d)

    # --- bump version + date du jour
    f = d["fiche"]
    f["version"] = bump_version(f.get("version", "v1.0"))
    f["date"] = today_fr()
    return changes, d


def recompute_dp(d):
    """Recale les ΔP de référence (classes.*.dp et dimensions[].dp) sur le nouveau polynôme."""
    dnom = d.get("debit_nom", 3400)
    if "classes" in d:
        low = d["classes"].get("low")
        if low and "poly" in low:
            co48 = low["poly"].get("48")
            if co48:
                cot = (co48["a"], co48["b"], co48["c"])
                low["dp"] = round(poly_at(cot, dnom / 3600 / AREA))
                if d["classes"].get("high"):
                    d["classes"]["high"]["dp"] = low["dp"]
        for dim in d.get("dimensions", []):
            ep = str(dim.get("P"))
            poly = low["poly"].get(ep) or low["poly"].get("48")
            if poly:
                area = dim["L"] * dim["H"] / 1_000_000
                v = dim["debit"] / 3600 / area
                dim["dp"] = round(poly_at((poly["a"], poly["b"], poly["c"]), v))


def bump_version(v):
    m = re.match(r"v(\d+)\.(\d+)", str(v))
    if not m:
        return "v1.1"
    return f"v{m.group(1)}.{int(m.group(2)) + 1}"


def today_fr():
    import datetime
    return datetime.date.today().strftime("%d/%m/%Y")


# ----------------------------------------------------------------- main ----
def main():
    if not os.path.exists(PDC):
        print(f"❌ Source introuvable : {PDC}")
        sys.exit(1)
    lookup = read_pdc()
    print(f"📊 DONNEES_PDC : {len(lookup)} courbes lues (ΔP → polynôme recalculé)\n")

    changed_slugs = []
    for path in sorted(glob.glob(os.path.join(PRODUITS, "*.json"))):
        if os.path.basename(path) == "_gabarit_ref.json":
            continue
        changes, newd = update_one(path, lookup)
        if changes:
            nom = json.load(open(path, encoding="utf-8")).get("nom", os.path.basename(path))
            print(f"🔧 {nom} ({len(changes)} changement(s)) :")
            for c in changes:
                print(f"     • {c}")
            if APPLY:
                print(f"     → version {newd['fiche']['version']} · {newd['fiche']['date']}")
                json.dump(newd, open(path, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
                open(path, "a", encoding="utf-8").write("\n")
                changed_slugs.append(newd["slug"])
            print()

    if not changed_slugs and not any(True for _ in []):
        if APPLY:
            print("✅ Rien à mettre à jour : tout est déjà cohérent avec DONNEES_PDC.")
        else:
            pass

    if APPLY and changed_slugs:
        print("♻️  Régénération des fiches modifiées (courbe + calculateur + synchro site)…")
        for slug in changed_slugs:
            subprocess.run([sys.executable, os.path.join(ROOT, "generer.py"),
                            os.path.join("produits", slug + ".json")], cwd=ROOT)
        print(f"\n✅ {len(changed_slugs)} fiche(s) mise(s) à jour de bout en bout.")
    elif not APPLY:
        print("— APERÇU (dry-run). Rien n'a été écrit.")
        print("  Lancez  python3 maj_fiches.py --apply  pour appliquer.")


if __name__ == "__main__":
    main()
